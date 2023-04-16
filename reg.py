import customtkinter
from customtkinter import *
from tkinter import *
from datetime import date
from datetime import *
from tkinter import filedialog,messagebox
from PIL import Image
import os
import openpyxl
from openpyxl import Workbook,load_workbook
import pathlib

customtkinter.set_appearance_mode("Light")


background = "#06283D"
framebg = "#EDEDED"
framefg = "#06283D"


root = customtkinter.CTk()
root.title("Student Registration System")
root.geometry("1250x700+210+100")
root.config(bg=background)
root.resizable(0,0)


file = pathlib.Path("Student_data.xlsx")
if file.exists():
    pass
else:
    file=Workbook()
    sheet = file.active
    sheet["A1"]="Registration No. "
    sheet["B1"]="Name"
    sheet["C1"]="Class"
    sheet["D1"]="Gender"
    sheet["E1"]="DOB"
    sheet["F1"]="Date of Registration"
    sheet["G1"]="Religion"
    sheet["H1"]="Skill"
    sheet["I1"]="Father Name"
    sheet["J1"]="Mother Name"
    sheet["K1"]="Father's Occupation"
    sheet["L1"]="Mother's Occupation"

    file.save("Student_data.xlsx")


#######gender#########
def selection():
    global gender
    value = radio.get()
    if value==1:
        gender="Male"
    else:
        gender="Female"


#######Exit###########   
def Exit():
    root.destroy()


############show image############
def showimage():
    global filename
    filename = filedialog.askopenfilename(initialdir=os.getcwd(),
                                          title="Select image file",filetypes=[('Images','*.jpg *.jpeg *.png')])
    
    
    your_image = customtkinter.CTkImage(light_image=Image.open(os.path.join(filename)), size=(200,200))
    lbl.configure(image=your_image)
    lbl.image=your_image

##########Registration NO.############

def registration_no():
    file = openpyxl.load_workbook('Student_data.xlsx')
    sheet=file.active
    row = sheet.max_row

    max_row_value = sheet.cell(row=row,column=1).value

    try:
        registration.set(max_row_value+1)

    except:
        registration.set("1")

########Clear#############

def Clear():
    global your_image
    global img
    Name.set('')
    DOB.set('')
    Religion.set('')
    skill.set('')
    F_name.set('')
    Mothers_name.set('')
    F_occupation.set('')
    Mothers_occupation.set('')
    Class.set("Select Class")
    

    registration_no()

    image = customtkinter.CTkImage(light_image=Image.open('C:\\Users\\Administrator\\Desktop\\My Python Projects\\upload photo.png'),size=(200,200))    
    lbl.configure(image=image)
    lbl.image=image

    img=""


#########save###########

def save():
    
    
    
    
    
    
    
    
    
    R1=registration.get()
    N1=Name.get()
    C1=Class.get()
    try:
        G1=gender
    except:
        messagebox.showerror("error","Select Gender!")
    
    D2=DOB.get()
    D1=date.get()
    Re1=Religion.get()
    S1=skill.get()
    F_name.get()
    Mothers_name.get()
    F1=F_occupation.get()
    M1=Mothers_occupation.get()

    if N1=="" or C1=="Select Class" or D2=="" or S1=="" or F_name=="" or Mothers_name=="" or F1=="" or M1=="" or D1=="" or Re1=="":
        messagebox.showerror("error","Few data is missing")

    else:
        file=openpyxl.load_workbook('Student_data.xlsx')
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=R1)
        sheet.cell(column=2,row=sheet.max_row+1,value=N1)
        sheet.cell(column=3,row=sheet.max_row+1,value=C1)
        sheet.cell(column=4,row=sheet.max_row+1,value=G1)
        sheet.cell(column=5,row=sheet.max_row+1,value=D2)
        sheet.cell(column=6,row=sheet.max_row+1,value=D1)
        sheet.cell(column=7,row=sheet.max_row+1,value=Re1)
        sheet.cell(column=8,row=sheet.max_row+1,value=S1)
        sheet.cell(column=9,row=sheet.max_row+1,value=F_name)
        sheet.cell(column=10,row=sheet.max_row+1,value=Mothers_name)
        sheet.cell(column=11,row=sheet.max_row+1,value=F1)
        sheet.cell(column=12,row=sheet.max_row+1,value=M1)
        file.save(r'Student_data.xlsx')
        





#Top Frame
customtkinter.CTkLabel(root,text="Email: datonaxucrishvili65@gmail.com",width=10,height=50,bg_color="#f0687c",anchor="e").pack(side=TOP,fill=X)
customtkinter.CTkLabel(root,text="STUDENT REGISTRATION",width=10,height=45,bg_color="#c36464",font=("arial",20,"bold"),).pack(side=TOP,fill=X)

search = StringVar()
customtkinter.CTkEntry(root,textvariable=search,width=200,bg_color="#c36464",font=("arial",20)).place(x=760,y=57)
image_icon4 = CTkImage(light_image=Image.open('C:\\Users\\Administrator\\Desktop\\My Python Projects\\Layer 4.png'))
srch = CTkButton(root,text="Search",image=image_icon4,compound=LEFT,width=123,bg_color="#c36464",font=("arial",14,"bold"))
srch.place(x=970,y=57)



image_icon3 = CTkImage(light_image=Image.open('C:\\Users\\Administrator\\Desktop\\My Python Projects\\search.png'))
update_button = CTkButton(root,image=image_icon3,bg_color="#c36464",width=40,height=35,text="")
update_button.place(x=40,y=58)

#reg and data

CTkLabel(root,text="Registration No: ",font=("arial",13),bg_color=background,text_color="white").place(x=30,y=150)
CTkLabel(root,text="Date: ",font=("arial",15),bg_color=background,text_color="white").place(x=500,y=150)

registration = StringVar()
date = StringVar()
today = StringVar()

reg_entry = CTkEntry(root,textvariable=registration,width=120,height=27,bg_color=background,font=("arial",15,"bold"))
reg_entry.place(x=125,y=152)


registration_no()

today = datetime.today()
d1 = today.strftime("%d %m %y")
data_entry = CTkEntry(root,textvariable=date,width=120,height=27,bg_color=background,font=("arial",15,"bold"))
data_entry.place(x=550,y=150)





date.set(d1)

#std details

obj = LabelFrame(root,text="Student's Details",font=20,width=900,bg=framebg,fg=framefg,height=250,relief=GROOVE)
obj.place(x=30,y=200)

CTkLabel(obj,text="Full Name: ",font=("arial",13),bg_color=framebg,).place(x=30,y=50)
CTkLabel(obj,text="Date of Birth: ",font=("arial",13),bg_color=framebg,).place(x=30,y=100)
CTkLabel(obj,text="Gender:  ",font=("arial",13),bg_color=framebg,).place(x=30,y=150)

CTkLabel(obj,text="Class: ",font=("arial",13),bg_color=framebg,).place(x=500,y=50)
CTkLabel(obj,text="Religion: ",font=("arial",13),bg_color=framebg,).place(x=500,y=100)
CTkLabel(obj,text="Skills: ",font=("arial",13),bg_color=framebg,).place(x=500,y=150)


Name =StringVar()
name_entry = CTkEntry(obj,textvariable=Name,width=130,font=("arial",15))
name_entry.place(x=160,y=50)

DOB =StringVar()
DOB_entry = CTkEntry(obj,textvariable=DOB,width=130,font=("arial",15))
DOB_entry.place(x=160,y=100)


radio = IntVar()
R1 = CTkRadioButton(obj,text="Male",variable=radio,value=1,bg_color=framebg,command=selection)
R1.place(x=150,y=150)

R2 = CTkRadioButton(obj,text="Female",variable=radio,value=2,bg_color=framebg,command=selection)
R2.place(x=250,y=150)

Religion = StringVar()
religion_entry = CTkEntry(obj,textvariable=Religion,width=130,font=("arial",15))
religion_entry.place(x=630,y=100)


skill =StringVar()
skill_entry = CTkEntry(obj,textvariable=skill,width=130,font=("arial",15))
skill_entry.place(x=630,y=150)

Class = CTkComboBox(obj,values=["1","2","3","4","5","6","7","8","9","10","11","12"],font=("Roboto",15),width=150)
Class.place(x=630,y=50)
Class.set("Select Class")


#parents details
obj2 = LabelFrame(root,text="Parent's Details",font=20,width=900,bg=framebg,fg=framefg,height=220,relief=GROOVE)
obj2.place(x=30,y=470)

CTkLabel(obj2,text="Father's Name:",font=("arial",13),bg_color=framebg).place(x=30,y=50)
CTkLabel(obj2,text="Occupation:",font=("arial",13),bg_color=framebg).place(x=30,y=100)

F_name =StringVar()
f_entry = CTkEntry(obj2,textvariable=F_name,width=130,font=("arial",15))
f_entry.place(x=160,y=50)

F_occupation =StringVar()
fo_entry = CTkEntry(obj2,textvariable=F_occupation,width=130,font=("arial",15))
fo_entry.place(x=160,y=100)

CTkLabel(obj2,text="Mother's Name:",font=("arial",15),bg_color=framebg).place(x=500,y=50)
CTkLabel(obj2,text="Occupation:",font=("arial",15),bg_color=framebg).place(x=500,y=100)

Mothers_name =StringVar()
m_entry = CTkEntry(obj2,textvariable=Mothers_name,width=130,font=("arial",15))
m_entry.place(x=630,y=50)

Mothers_occupation =StringVar()
mo_entry = CTkEntry(obj2,textvariable=Mothers_occupation,width=130,font=("arial",15))
mo_entry.place(x=630,y=100)

#image
f=CTkFrame(root,bg_color="black",width=200,height=200)
f.place(x=1000,y=150)

img = CTkImage(light_image=Image.open('C:\\Users\\Administrator\\Desktop\\My Python Projects\\upload photo.png'),size=(200,200))
lbl = CTkLabel(f,bg_color="black",image=img,text="")
lbl.place(x=0,y=0)

#button

CTkButton(root,text="Upload",width=200,height=50,font=("arial",15,"bold"),text_color="black",fg_color="lightblue",command=showimage).place(x=1000,y=370)

save_button=CTkButton(root,text="Save",width=200,height=50,font=("arial",15,"bold"),text_color="black",fg_color="lightgreen",command=save)
save_button.place(x=1000,y=450)

CTkButton(root,text="Reset",width=200,height=50,font=("arial",15,"bold"),text_color="black",fg_color="lightpink",command=Clear).place(x=1000,y=530)

CTkButton(root,text="Exit",width=200,height=50,font=("arial",15,"bold"),text_color="black",fg_color="grey",command=Exit).place(x=1000,y=610)






root.mainloop()